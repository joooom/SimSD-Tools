#!/usr/bin/env python3
import json
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_duration(seconds):
    if seconds is None:
        return "00:00"
    try:
        seconds = int(seconds)
    except ValueError:
        return "00:00"
    m = seconds // 60
    s = seconds % 60
    return f"{m:02d}:{s:02d}"

def main():
    if len(sys.argv) < 2:
        print("Uso: python json_to_excel.py <caminho_do_json> [caminho_do_excel]")
        sys.exit(1)

    json_path = sys.argv[1]
    if not os.path.exists(json_path):
        print(f"Erro: Arquivo JSON não encontrado: {json_path}")
        sys.exit(1)

    # Load JSON
    with open(json_path, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError as e:
            print(f"Erro ao decodificar JSON: {e}")
            sys.exit(1)

    # Resolve output path
    if len(sys.argv) >= 3:
        excel_path = sys.argv[2]
    else:
        # Default filename based on committee name and date
        committee_name = data.get("config", {}).get("committee", "comite").lower().replace(" ", "-")
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        excel_path = f"simsd-sessao-{committee_name}-{timestamp}.xlsx"

    # Create openpyxl Workbook
    wb = Workbook()
    
    # Setup Styles
    FONT_FAMILY = "Segoe UI"
    
    # Palette matching the app's premium warm dark red theme
    COLOR_PRIMARY = "8B1A1A"      # Warm dark red for main headers
    COLOR_SECONDARY = "FDF2F2"    # Very light red for zebra striping
    COLOR_BORDER = "E0DBD4"       # Clean soft beige-gray for borders
    COLOR_MUTED_BG = "F5F3EF"     # Soft light gray for section titles/meta
    COLOR_APPROVED = "EDF7F1"     # Light green for approved items
    COLOR_REJECTED = "FDF2F2"     # Light red for rejected items
    COLOR_PENDING = "FFF9E6"      # Light yellow for pending items
    
    # Font definitions
    title_font = Font(name=FONT_FAMILY, size=16, bold=True, color=COLOR_PRIMARY)
    section_font = Font(name=FONT_FAMILY, size=12, bold=True, color="1A1714")
    header_font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=FONT_FAMILY, size=10, color="1A1714")
    bold_data_font = Font(name=FONT_FAMILY, size=10, bold=True, color="1A1714")
    meta_label_font = Font(name=FONT_FAMILY, size=10, bold=True, color="6A635C")
    meta_val_font = Font(name=FONT_FAMILY, size=10, color="1A1714")
    
    # Fills
    header_fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    zebra_fill = PatternFill(start_color=COLOR_SECONDARY, end_color=COLOR_SECONDARY, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    meta_fill = PatternFill(start_color=COLOR_MUTED_BG, end_color=COLOR_MUTED_BG, fill_type="solid")
    
    fill_ok = PatternFill(start_color="EDF7F1", end_color="EDF7F1", fill_type="solid")
    fill_no = PatternFill(start_color="FDF2F2", end_color="FDF2F2", fill_type="solid")
    fill_warn = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

    # Alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Borders
    border_thin = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER)
    )

    # -------------------------------------------------------------
    # TAB 1: RESUMO (Overview)
    # -------------------------------------------------------------
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo.views.sheetView[0].showGridLines = True
    
    # Title
    ws_resumo["A1"] = "SimSD — Relatório Geral da Sessão"
    ws_resumo["A1"].font = title_font
    
    # General Info
    ws_resumo["A3"] = "Configurações Gerais"
    ws_resumo["A3"].font = section_font
    
    configs = [
        ("Conferência", data.get("config", {}).get("conference", "SimSD")),
        ("Comitê", data.get("config", {}).get("committee", "—")),
        ("Sessão", data.get("config", {}).get("session", "—")),
        ("Tema / Agenda", data.get("agenda", "—")),
        ("Tipo de Comitê", "Câmara dos Deputados" if data.get("committeeType") == "camara" else "Modelo ONU"),
        ("Exportado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
    ]
    
    for idx, (lbl, val) in enumerate(configs, start=4):
        ws_resumo[f"A{idx}"] = lbl
        ws_resumo[f"A{idx}"].font = meta_label_font
        ws_resumo[f"A{idx}"].fill = meta_fill
        ws_resumo[f"A{idx}"].border = border_thin
        
        ws_resumo[f"B{idx}"] = val
        ws_resumo[f"B{idx}"].font = meta_val_font
        ws_resumo[f"B{idx}"].alignment = align_left
        ws_resumo[f"B{idx}"].border = border_thin

    # Metrics
    start_row_metrics = 12
    ws_resumo[f"A{start_row_metrics}"] = "Métricas Rápidas"
    ws_resumo[f"A{start_row_metrics}"].font = section_font
    
    countries_list = data.get("committeeCountries", [])
    presence_map = data.get("presence", {})
    history_list = data.get("history", [])
    motions_list = data.get("motions", [])
    vote_history = data.get("voteHistory", [])
    
    total_delegates = len(countries_list)
    present_delegates = sum(1 for c in countries_list if presence_map.get(c.get("c")) in ("presente", "presente-votante"))
    total_speeches = sum(data.get("speeches", {}).values())
    total_speak_time = sum(data.get("speakTime", {}).values())
    motions_approved = sum(1 for m in motions_list if m.get("status") == "approved")
    
    metrics = [
        ("Total de Delegados", total_delegates),
        ("Delegados Presentes", present_delegates),
        ("Discursos na GSL", total_speeches),
        ("Tempo Total Falado", format_duration(total_speak_time)),
        ("Moções Aprovadas", motions_approved),
        ("Votações Realizadas", len(vote_history)),
    ]
    
    for idx, (lbl, val) in enumerate(metrics, start=start_row_metrics+1):
        ws_resumo[f"A{idx}"] = lbl
        ws_resumo[f"A{idx}"].font = meta_label_font
        ws_resumo[f"A{idx}"].fill = meta_fill
        ws_resumo[f"A{idx}"].border = border_thin
        
        ws_resumo[f"B{idx}"] = val
        ws_resumo[f"B{idx}"].font = bold_data_font
        ws_resumo[f"B{idx}"].alignment = align_left
        ws_resumo[f"B{idx}"].border = border_thin

    # Voting Rules / Quorums
    voting_present = sum(1 for c in countries_list if presence_map.get(c.get("c")) == "presente-votante" and c.get("voto") is not False)
    ms = (voting_present // 2) + 1 if voting_present > 0 else 0
    import math
    mq = math.ceil(voting_present * 2 / 3) if voting_present > 0 else 0
    
    start_row_quorums = 21
    ws_resumo[f"A{start_row_quorums}"] = "Quóruns de Votação (baseado em Presentes Votantes)"
    ws_resumo[f"A{start_row_quorums}"].font = section_font
    
    quorums = [
        ("Delegações Presentes e Votantes", voting_present),
        ("Maioria Simples (50% + 1)", ms),
        ("Maioria Qualificada (2/3)", mq),
    ]
    
    for idx, (lbl, val) in enumerate(quorums, start=start_row_quorums+1):
        ws_resumo[f"A{idx}"] = lbl
        ws_resumo[f"A{idx}"].font = meta_label_font
        ws_resumo[f"A{idx}"].fill = meta_fill
        ws_resumo[f"A{idx}"].border = border_thin
        
        ws_resumo[f"B{idx}"] = val
        ws_resumo[f"B{idx}"].font = bold_data_font
        ws_resumo[f"B{idx}"].alignment = align_left
        ws_resumo[f"B{idx}"].border = border_thin

    # -------------------------------------------------------------
    # TAB 2: LISTA DE PRESENÇA (Attendance & Stats)
    # -------------------------------------------------------------
    ws_presence = wb.create_sheet(title="Lista de Presença")
    ws_presence.views.sheetView[0].showGridLines = True
    
    ws_presence["A1"] = "Lista de Presença e Métricas de Oradores"
    ws_presence["A1"].font = title_font
    
    headers_presence = [
        "Participante / Delegação", "Bandeira", "Categoria", "Subtítulo / Filiação",
        "Região", "Status de Presença", "Direito a Voto", "Discursos (GSL)",
        "Tempo Total Falado (MM:SS)", "Tempo Total Falado (segundos)"
    ]
    
    for col_idx, h in enumerate(headers_presence, start=1):
        cell = ws_presence.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx != 1 and col_idx != 4 else align_left
        cell.border = border_thin
        
    presence_translations = {
        "presente": "Presente",
        "presente-votante": "Presente e Votante",
        "ausente": "Ausente"
    }
    
    for r_idx, c in enumerate(countries_list, start=4):
        c_name = c.get("c", "—")
        flag = c.get("f", "")
        category = c.get("cat") or "—"
        subtitle = c.get("sub") or "—"
        region = c.get("r", "—").capitalize()
        status_raw = presence_map.get(c_name, "ausente")
        status = presence_translations.get(status_raw, "Ausente")
        can_vote = "Sim" if c.get("voto") is not False else "Não"
        speech_count = data.get("speeches", {}).get(c_name, 0)
        speak_sec = data.get("speakTime", {}).get(c_name, 0)
        speak_fmt = format_duration(speak_sec)
        
        row_values = [
            c_name, flag, category, subtitle, region, status, can_vote,
            speech_count, speak_fmt, speak_sec
        ]
        
        is_zebra = (r_idx % 2 == 0)
        row_fill = zebra_fill if is_zebra else white_fill
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws_presence.cell(row=r_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = border_thin
            
            # Formatting
            if col_idx in (2, 5, 7):
                cell.alignment = align_center
            elif col_idx in (6,):
                cell.alignment = align_center
                if status_raw == "presente-votante":
                    cell.fill = fill_ok
                    cell.font = bold_data_font
                elif status_raw == "presente":
                    cell.fill = fill_warn
                else:
                    cell.fill = fill_no
            elif col_idx in (8, 10):
                cell.alignment = align_right
                cell.number_format = '#,##0'
            elif col_idx == 9:
                cell.alignment = align_right
            else:
                cell.alignment = align_left

    # -------------------------------------------------------------
    # TAB 3: HISTÓRICO GSL (GSL Speech History)
    # -------------------------------------------------------------
    ws_gsl = wb.create_sheet(title="Histórico GSL")
    ws_gsl.views.sheetView[0].showGridLines = True
    
    ws_gsl["A1"] = "Histórico Cronológico Geral de Oradores (GSL)"
    ws_gsl["A1"].font = title_font
    
    headers_gsl = [
        "Ordem", "Hora de Registro", "Orador", "Bandeira",
        "Tempo Falado (MM:SS)", "Tempo Falado (segundos)", "Cessão / Recebimento"
    ]
    
    for col_idx, h in enumerate(headers_gsl, start=1):
        cell = ws_gsl.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx not in (3, 7) else align_left
        cell.border = border_thin

    # History is saved newest first, so we reverse it to display chronological order
    chrono_history = list(reversed(history_list))
    
    for r_idx, h in enumerate(chrono_history, start=4):
        ordem = r_idx - 3
        timestamp = h.get("t", "—")
        orador = h.get("c", "—")
        flag = h.get("f", "")
        sec = h.get("sec", 0)
        duration = format_duration(sec)
        
        # Build Yield / Received description
        obs_parts = []
        if h.get("received"):
            obs_parts.append(f"Recebeu de {h['received']}")
        yielded = h.get("yielded")
        if yielded == "chair":
            obs_parts.append("Cedeu ao Chair")
        elif yielded:
            obs_parts.append(f"Cedeu a {yielded}")
            
        obs = " · ".join(obs_parts) if obs_parts else "—"
        
        row_values = [ordem, timestamp, orador, flag, duration, sec, obs]
        is_zebra = (r_idx % 2 == 0)
        row_fill = zebra_fill if is_zebra else white_fill
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws_gsl.cell(row=r_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = border_thin
            
            if col_idx in (1, 2, 4):
                cell.alignment = align_center
            elif col_idx in (5, 6):
                cell.alignment = align_right
                if col_idx == 6:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = align_left

    # -------------------------------------------------------------
    # TAB 4: MOÇÕES (Motions History)
    # -------------------------------------------------------------
    ws_motions = wb.create_sheet(title="Moções")
    ws_motions.views.sheetView[0].showGridLines = True
    
    ws_motions["A1"] = "Histórico de Moções Propostas"
    ws_motions["A1"].font = title_font
    
    headers_motions = [
        "Hora", "Proponente", "Tipo de Moção", "Duração (min)",
        "Tempo Orador (seg)", "Status", "Decisão"
    ]
    
    for col_idx, h in enumerate(headers_motions, start=1):
        cell = ws_motions.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx != 2 else align_left
        cell.border = border_thin
        
    motion_type_map = {
        "mod": "Sessão Moderada",
        "unmod": "Sessão Não-Moderada",
        "vote": "Votação de Documento",
        "recess": "Recesso",
        "other": "Outra"
    }
    
    motion_status_map = {
        "pending": "Pendente",
        "approved": "Aprovada",
        "rejected": "Rejeitada"
    }
    
    # Motions list is chronologically in reverse order in state, let's reverse it to be chronological
    chrono_motions = list(reversed(motions_list))
    
    for r_idx, m in enumerate(chrono_motions, start=4):
        timestamp = m.get("ts", "—")
        proponent = m.get("prop", "—")
        m_type = motion_type_map.get(m.get("type"), m.get("type", "—"))
        dur = m.get("dur")
        try:
            dur = int(dur) if dur else ""
        except ValueError:
            pass
        spk = m.get("spk")
        try:
            spk = int(spk) if spk else ""
        except ValueError:
            pass
        status_raw = m.get("status", "pending")
        status = motion_status_map.get(status_raw, status_raw)
        
        row_values = [timestamp, proponent, m_type, dur, spk, status_raw.capitalize(), status]
        is_zebra = (r_idx % 2 == 0)
        row_fill = zebra_fill if is_zebra else white_fill
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws_motions.cell(row=r_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = border_thin
            
            if col_idx in (1, 3, 6, 7):
                cell.alignment = align_center
                if col_idx == 7:
                    if status_raw == "approved":
                        cell.fill = fill_ok
                        cell.font = bold_data_font
                    elif status_raw == "rejected":
                        cell.fill = fill_no
                    else:
                        cell.fill = fill_warn
            elif col_idx in (4, 5):
                cell.alignment = align_right
                if isinstance(val, int):
                    cell.number_format = '#,##0'
            else:
                cell.alignment = align_left

    # -------------------------------------------------------------
    # TAB 5: HISTÓRICO DE VOTAÇÕES (Votes Overview & Breakdown)
    # -------------------------------------------------------------
    ws_votes = wb.create_sheet(title="Histórico de Votações")
    ws_votes.views.sheetView[0].showGridLines = True
    
    ws_votes["A1"] = "Histórico de Votações Concluídas"
    ws_votes["A1"].font = title_font
    
    # 1. Master list of votes
    ws_votes["A3"] = "Resumo das Votações"
    ws_votes["A3"].font = section_font
    
    headers_votes_master = [
        "ID/Ordem", "Título / Descrição", "Tipo", "Hora", "Regra de Maioria",
        "Votos Necessários", "Favoráveis Calculados", "Resultado", "Tally (Favor)",
        "Tally (Fav c/ Dir)", "Tally (Contra)", "Tally (Con c/ Dir)", "Tally (Abstenções)"
    ]
    
    for col_idx, h in enumerate(headers_votes_master, start=1):
        cell = ws_votes.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx not in (2,) else align_left
        cell.border = border_thin
        
    chrono_votes = list(reversed(vote_history))
    
    vote_type_map = {
        "procedimental": "Procedimental",
        "substancial": "Substancial"
    }
    
    majority_map = {
        "simples": "Simples (50%+1)",
        "qualificada": "Qualificada (2/3)",
        "csnu": "CSNU (9 votos)",
        "consenso": "Consenso"
    }
    
    for r_idx, v in enumerate(chrono_votes, start=6):
        label = v.get("label", "—")
        v_type = vote_type_map.get(v.get("type"), v.get("type", "—"))
        timestamp = v.get("t", "—")
        maj_rule = majority_map.get(v.get("majority"), v.get("majority", "—"))
        req_quorum = v.get("maj", 0)
        fav_total = v.get("totalFavor", 0)
        approved_bool = v.get("approved", False)
        result = "Aprovada" if approved_bool else "Não Aprovada"
        
        tally = v.get("tally", {})
        t_fav = tally.get("fav", 0)
        t_fdr = tally.get("fdr", 0)
        t_con = tally.get("con", 0)
        t_cdr = tally.get("cdr", 0)
        t_abs = tally.get("abs", 0)
        
        row_values = [
            r_idx - 5, label, v_type, timestamp, maj_rule, req_quorum, fav_total,
            result, t_fav, t_fdr, t_con, t_cdr, t_abs
        ]
        
        is_zebra = (r_idx % 2 == 0)
        row_fill = zebra_fill if is_zebra else white_fill
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws_votes.cell(row=r_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = border_thin
            
            if col_idx in (1, 3, 4, 5, 8):
                cell.alignment = align_center
                if col_idx == 8:
                    if approved_bool:
                        cell.fill = fill_ok
                        cell.font = bold_data_font
                    else:
                        cell.fill = fill_no
            elif col_idx >= 6:
                cell.alignment = align_right
                cell.number_format = '#,##0'
            else:
                cell.alignment = align_left

    # 2. Detailed matrix of individual votes per delegation
    start_row_matrix = len(chrono_votes) + 9
    ws_votes[f"A{start_row_matrix}"] = "Detalhamento de Votos Individuais por Delegação"
    ws_votes[f"A{start_row_matrix}"].font = section_font
    
    # We construct a table where rows are delegations and columns are the vote descriptions
    if len(chrono_votes) > 0:
        headers_matrix = ["Delegação / Participante"] + [v.get("label", f"Votação {idx+1}") for idx, v in enumerate(chrono_votes)]
        
        for col_idx, h in enumerate(headers_matrix, start=1):
            cell = ws_votes.cell(row=start_row_matrix + 2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center if col_idx != 1 else align_left
            cell.border = border_thin
            
        vote_val_translations = {
            "fav": "A Favor",
            "fdr": "A Favor c/ Dir.",
            "con": "Contra",
            "cdr": "Contra c/ Dir.",
            "abs": "Abstenção"
        }
        
        vote_val_fills = {
            "fav": fill_ok,
            "fdr": fill_ok,
            "con": fill_no,
            "cdr": fill_no,
            "abs": fill_warn
        }
        
        for r_idx, c in enumerate(countries_list, start=start_row_matrix + 3):
            c_name = c.get("c", "—")
            is_zebra = (r_idx % 2 == 0)
            row_fill = zebra_fill if is_zebra else white_fill
            
            # Delegation column
            cell = ws_votes.cell(row=r_idx, column=1, value=c_name)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = border_thin
            cell.alignment = align_left
            
            # Individual vote columns
            for v_idx, v in enumerate(chrono_votes, start=2):
                v_details = v.get("detail", [])
                delegation_vote = next((d.get("v") for d in v_details if d.get("c") == c_name), None)
                
                txt_vote = vote_val_translations.get(delegation_vote, "—")
                cell = ws_votes.cell(row=r_idx, column=v_idx, value=txt_vote)
                cell.font = data_font
                cell.fill = row_fill
                
                # Apply custom cell colors/styles for cast votes
                if delegation_vote in vote_val_fills:
                    cell.fill = vote_val_fills[delegation_vote]
                    if "fav" in delegation_vote or "con" in delegation_vote:
                        cell.font = bold_data_font
                        
                cell.border = border_thin
                cell.alignment = align_center
    else:
        ws_votes.cell(row=start_row_matrix + 2, column=1, value="Nenhuma votação registrada nesta sessão.").font = data_font

    # -------------------------------------------------------------
    # POST-PROCESSING: AUTO-FIT COLUMN WIDTHS & GRIDLINES
    # -------------------------------------------------------------
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Inspect first 100 rows to avoid checking massive sheets fully
            for cell in col[:100]:
                val = str(cell.value or '')
                # Adjust for potential line breaks
                val_len = max(len(sub) for sub in val.split('\n')) if val else 0
                
                # Boost size for title banner / merged rows to prevent huge column stretches
                if cell.row == 1 or cell.row == start_row_matrix:
                    continue
                    
                # Flag emojis are wide, let's treat them as small lengths
                if cell.column == 2 and ws.title in ("Lista de Presença", "Histórico GSL"):
                    val_len = 3
                    
                if val_len > max_len:
                    max_len = val_len
                    
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save Excel Workbook
    try:
        wb.save(excel_path)
        print(f"Planilha Excel criada com sucesso em: {excel_path}")
    except IOError as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
